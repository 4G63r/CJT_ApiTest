#!/usr/bin/env python

# @Author: songxiao
# @Time: 2019-09-25 18:49


from openpyxl import load_workbook
from conf import filePath


class OperaExcel:
    """操作excel"""

    def __init__(self, file_name=None, sheet_name=None):
        if file_name is None:
            self.file_name = filePath.testcase_abspath
            self.sheet_name = 'testcase'
        else:
            self.file_name = file_name
            self.sheet_name = sheet_name

        self.wb = load_workbook(self.file_name)
        self.ws = self.wb[self.sheet_name]

    @property
    def max_row(self):
        """最大行数"""
        return self.ws.max_row

    def get_all_sheets(self):
        """
        获得所有的sheet名称
        :return: ['Sheet1', 'Sheet2']
        """
        return self.wb.sheetnames

    def is_have_sheet(self, sheet_name=None):
        """
        判断是否包含sheet
        :return:
        """
        if sheet_name is None:
            sheet_name = self.sheet_name
        if sheet_name in self.get_all_sheets():
            return True
        return False

    def get_cell_value(self, row, col):
        """获取单元格内容"""
        value = self.ws.cell(row, col).value
        return value

    def read_excel(self):
        col = self.ws.max_column
        row = self.ws.max_row
        test_data = ({self.ws.cell(2, j).value: self.ws.cell(i, j).value for j in range(1, col + 1)} for i in
                     range(3, row + 1))
        return test_data

    def save_excel(self):
        """写入操作后最后需要保存excel"""
        try:
            self.wb.save(self.file_name)
        except Exception as e:
            print(e)

    def load_excel_data(self):
        """加载excel数据，以第一列数据为键，保存各行数据"""

        try:
            datas = {}
            for i in self.get_all_sheets():
                sheet = self.wb[i]
                for row in range(1, sheet.max_row):
                    # 每一行的数据在row_data 数组里
                    row_data = self.get_row_data(row)
                    datas[row_data[0]] = row_data
            return datas
        except Exception as e:
            raise ValueError("加载excel数据出现异常,文件：%s，异常：%s" % (self.file_name, e))

    def get_row_data(self, row):
        row_datas = []
        for cell in list(self.ws.rows)[row]:
            row_datas.append(cell.value)
        return row_datas

    def get_excel_data_by_key(self, key):
        """根据关键列data获取sheet里行数据"""
        datas = self.load_excel_data()
        return datas.get(key)

    def all_case_data(self):
        """根据是否执行获取全部cases列表"""
        all_case_datas = []
        for row in range(2, self.max_row + 1):
            is_run = self.get_cell_value(row, 3)
            if is_run == 1:  # 执行
                all_case_datas.append(self.get_cell_value(row, 4))

        return all_case_datas
